"""Independently evaluate the exported audit workbook's supported Excel formulas."""
import sys, math, json
import openpyxl
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import range_boundaries
w=openpyxl.load_workbook(sys.argv[1]); cached=openpyxl.load_workbook(sys.argv[1],data_only=True)
memo={}
def flat(x):
    return [v for a in x for v in (a if isinstance(a,list) else [a])]
def ref(s,v):
    if '!' in v: s,v=v.rsplit('!',1);s=s.strip("'")
    v=v.replace('$','')
    if ':' in v:
        c1,r1,c2,r2=range_boundaries(v)
        return [cell(s,w[s].cell(r,c).coordinate) for r in range(r1,r2+1) for c in range(c1,c2+1)]
    return cell(s,v)
def cell(s,a):
    if (s,a) in memo:return memo[s,a]
    c=w[s][a];v=c.value
    if c.data_type=='f':v=evaluate(s,v)
    elif v is None:v=0
    memo[s,a]=v;return v
def evaluate(s,f):
    ts=[t for t in Tokenizer(f).items if t.type!='WHITE-SPACE'];i=0
    def expr(minp=0):
        nonlocal i
        t=ts[i];i+=1
        if t.type=='OPERATOR-PREFIX':v=expr(5);v=-v if t.value=='-' else v
        elif t.type=='PAREN' and t.subtype=='OPEN':v=expr();i+=1
        elif t.type=='FUNC' and t.subtype=='OPEN':
            name=t.value[:-1];args=[]
            while ts[i].subtype!='CLOSE':
                args.append(expr())
                if ts[i].type=='SEP':i+=1
                else:break
            i+=1;nums=[x for x in flat(args) if isinstance(x,(int,float))]
            if name=='SUM':v=sum(nums)
            elif name=='MAX':v=max(nums)
            elif name=='ABS':v=abs(args[0])
            elif name=='IF':v=args[1] if args[0] else args[2]
            elif name=='IFERROR':v=args[0]
            elif name=='COUNTIF':v=sum(x==args[1] for x in args[0])
            elif name=='SUMIF':v=sum(b for a,b in zip(args[0],args[2]) if a==args[1])
            elif name=='MATCH':v=args[1].index(args[0])+1
            elif name=='INDEX':v=args[0][int(args[1])-1]
            else:raise ValueError(name)
        elif t.subtype=='NUMBER':v=float(t.value)
        elif t.subtype=='TEXT':v=t.value[1:-1].replace('""','"')
        elif t.subtype=='RANGE':v=ref(s,t.value)
        else:raise ValueError((t.type,t.subtype,t.value))
        ops={'=':1,'<=':1,'>=':1,'<':1,'>':1,'+':2,'-':2,'*':3,'/':3,'^':4}
        while i<len(ts) and ts[i].type=='OPERATOR-INFIX' and ops.get(ts[i].value,0)>=minp:
            op=ts[i].value;p=ops[op];i+=1;b=expr(p+1)
            if op=='+':v+=b
            elif op=='-':v-=b
            elif op=='*':v*=b
            elif op=='/':v/=b
            elif op=='^':v=v**b
            elif op=='=':v=v==b
            elif op=='<=':v=v<=b
            elif op=='>=':v=v>=b
            elif op=='<':v=v<b
            elif op=='>':v=v>b
        return v
    return expr()
count=0
for s in w:
    for row in s:
        for c in row:
            if c.data_type!='f':continue
            actual=cell(s.title,c.coordinate);expected=cached[s.title][c.coordinate].value
            if isinstance(actual,(int,float)) and not isinstance(actual,bool):
                assert math.isclose(actual,expected,rel_tol=1e-10,abs_tol=1e-6),(s.title,c.coordinate,actual,expected)
            else:assert actual==expected,(s.title,c.coordinate,actual,expected)
            count+=1
print(json.dumps({'status':'PASS','sheets':len(w.sheetnames),'formulas_independently_evaluated':count}))
